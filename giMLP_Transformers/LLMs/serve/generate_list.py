

import gradio as gr
from model import ModelServe1, ModelServe2
import sys
import os
import argparse
import openpyxl

if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='training args.')
    parser.add_argument('--base_model', type=str, default='no-rep', help='Base Model Path.')
    parser.add_argument('--model_type', type=str, default='wp', help='Model Type')
    parser.add_argument('--input_file', type=str, default='../Input_List.txt', help='Input File')
    parser.add_argument('--output_file', type=str, default='', help='Output File')
    parser.add_argument('--load_8bit', type=bool, default=True, help='')
    parser.add_argument('--finetuned_weights', type=str, default='', help='')


    args = parser.parse_args()
    print('**************************  args  ******************************')
    print(f"arg is: {args}")
    print('**************************  args  ******************************')

    model1 = ModelServe1(load_8bit=False, model_type=args.model_type, base_model=args.base_model, finetuned_weights=args.finetuned_weights)
    model2 = ModelServe2(load_8bit=False, model_type=args.model_type, base_model=args.base_model, finetuned_weights=args.finetuned_weights)

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()


    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active


    # Using readline()
    file1 = open(args.input_file, 'r')
    count = 0


    while True:

        # Get next line from file
        inputs = file1.readline()
        if not inputs:
            break
        # print(inputs)
        # if line is empty
        # end of file is reached
        if "Topic" in inputs:
            topics = inputs[7:]
            print(topics)
            continue

        if "Input" in inputs:
            contents = inputs[7:]
            print(contents)
            count += 1

        instructions = "What suggestions or comments you can provide to address or alleviate  " + str(topics)
        outputs1 = model1.generate(instructions, contents)
        outputs2 = model2.generate(instructions, contents + " And tell me more knowledge.")
        print("Instruction{}: {}".format(count, instructions.strip()))
        print("Input{}: {}".format(count, contents.strip()))
        print("Output{}: {}".format(count, outputs1.strip()))
        print("Output{}: {}".format(count, outputs2.strip()))


        # Note: The first row or column integer
        # is 1, not 0. Cell object is created by
        # using sheet object's cell() method.

        # cc1 = sheet.cell(row=1, column=1)
        # cc1.value = "NO."
        #
        # cc2 = sheet.cell(row=1, column=2)
        # cc2.value = "Instructions"
        #
        # cc3 = sheet.cell(row=1, column=3)
        # cc3.value = "Inputs"
        #
        # cc4 = sheet.cell(row=1, column=4)
        # cc4.value = "Original"
        #
        # cc5 = sheet.cell(row=1, column=5)
        # cc5.value = "FineTuned"
        #
        #
        # c1 = sheet.cell(row=count+1, column=1)
        # c1.value = count
        #
        # c2 = sheet.cell(row=count+1, column=2)
        # c2.value = instructions
        #
        # c3 = sheet.cell(row=count+1, column=3)
        # c3.value = inputs
        #
        # c4 = sheet.cell(row=count+1, column=4)
        # c4.value = outputs1
        #
        # c5 = sheet.cell(row=count+1, column=5)
        # c5.value = outputs2


        # # Once have a Worksheet object, one can
        # # access a cell object by its name also.
        # # A2 means column = 1 & row = 2.
        # c3 = sheet['A2']
        # c3.value = "Welcome"
        #
        # # B2 means column = 2 & row = 2.
        # c4 = sheet['B2']
        # c4.value = "Everyone"





        cc1 = sheet.cell(row=1, column=1)
        cc1.value = "NO."

        cc3 = sheet.cell(row=1, column=2)
        cc3.value = "Topic"

        cc3 = sheet.cell(row=1, column=3)
        cc3.value = "Inputs"

        cc4 = sheet.cell(row=1, column=4)
        cc4.value = "Original"

        cc5 = sheet.cell(row=1, column=5)
        cc5.value = "FineTuned"

        cc6 = sheet.cell(row=1, column=6)
        cc6.value = "Readability-Original"

        cc6 = sheet.cell(row=1, column=7)
        cc6.value = "Readability-Finetuned"

        cc8 = sheet.cell(row=1, column=8)
        cc8.value = "Professional-Original"

        cc9 = sheet.cell(row=1, column=9)
        cc9.value = "Professional-Finetuned"

        cc10 = sheet.cell(row=1, column=10)
        cc10.value = "Match-Original"

        cc11 = sheet.cell(row=1, column=11)
        cc11.value = "Match-Finetuned"



        c1 = sheet.cell(row=count+1, column=1)
        c1.value = count

        c2 = sheet.cell(row=count+1, column=2)
        c2.value = topics

        c3 = sheet.cell(row=count+1, column=3)
        c3.value = contents

        c4 = sheet.cell(row=count+1, column=4)
        c4.value = outputs1

        c5 = sheet.cell(row=count+1, column=5)
        c5.value = outputs2


        wb.save(args.output_file)










